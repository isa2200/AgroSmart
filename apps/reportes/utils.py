import os
import io
import csv
from datetime import datetime, timedelta
from django.conf import settings
from django.template.loader import render_to_string
from django.http import HttpResponse
from django.db.models import Sum, Avg, Count, Q, F
from django.utils import timezone

# Importaciones opcionales para reportes avanzados
try:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import letter, A4
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch
    REPORTLAB_AVAILABLE = True
except ImportError:
    REPORTLAB_AVAILABLE = False

try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill
    from openpyxl.chart import BarChart, LineChart, Reference
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

from apps.aves.models import LoteAves, BitacoraDiaria

class GeneradorReportes:
    """
    Clase para generar reportes en diferentes formatos.
    """
    
    def __init__(self, tipo_reporte, parametros=None):
        self.tipo_reporte = tipo_reporte
        self.parametros = parametros or {}
        
    def obtener_datos_produccion(self):
        """
        Obtiene datos de producción según los parámetros especificados.
        """
        filtros_fecha = {}
        if 'fecha_inicio' in self.parametros:
            filtros_fecha['fecha__gte'] = self.parametros['fecha_inicio']
        if 'fecha_fin' in self.parametros:
            filtros_fecha['fecha__lte'] = self.parametros['fecha_fin']
        
        # Usar BitacoraDiaria como fuente principal
        produccion_lotes = BitacoraDiaria.objects.filter(**filtros_fecha).select_related('lote')
        datos = {}
        
        # Datos por lote
        datos['lotes'] = list(produccion_lotes.values(
            'lote__codigo',
            'lote__nombre_lote',
            'lote__linea',
            'produccion_aaa',
            'produccion_aa', 
            'produccion_a',
            'produccion_b',
            'produccion_c',
            'fecha'
        ))
        
        # Calcular totales de producción para cada registro
        for item in datos['lotes']:
            item['total_produccion'] = (
                (item['produccion_aaa'] or 0) +
                (item['produccion_aa'] or 0) +
                (item['produccion_a'] or 0) +
                (item['produccion_b'] or 0) +
                (item['produccion_c'] or 0)
            )
        
        # Resumen estadístico
        datos['resumen'] = {
            'total_huevos': sum(item['total_produccion'] for item in datos['lotes']),
            'promedio_huevos_lote': sum(item['total_produccion'] for item in datos['lotes']) / len(datos['lotes']) if datos['lotes'] else 0,
        }
        
        return datos
    
    def obtener_datos_produccion(self):
        """
        Obtiene datos de producción de aves según los parámetros especificados.
        """
        filtros_fecha = {}
        if 'fecha_inicio' in self.parametros:
            filtros_fecha['fecha__gte'] = self.parametros['fecha_inicio']
        if 'fecha_fin' in self.parametros:
            filtros_fecha['fecha__lte'] = self.parametros['fecha_fin']
        
        # Usar BitacoraDiaria como fuente principal
        produccion_aves = BitacoraDiaria.objects.filter(**filtros_fecha).select_related('lote')
        datos = {}
        
        # Datos por ave/lote
        datos['aves'] = list(produccion_aves.values(
            'lote__codigo',
            'lote__raza',
            'lote__linea',
            'produccion_aaa',
            'produccion_aa',
            'produccion_a', 
            'produccion_b',
            'produccion_c',
            'fecha'
        ))
        
        # Calcular totales de producción para cada registro
        for item in datos['aves']:
            item['total_produccion'] = (
                (item['produccion_aaa'] or 0) +
                (item['produccion_aa'] or 0) +
                (item['produccion_a'] or 0) +
                (item['produccion_b'] or 0) +
                (item['produccion_c'] or 0)
            )
        
        # Resumen estadístico
        datos['resumen'] = {
            'total_huevos': sum(item['total_produccion'] for item in datos['aves']),
            'promedio_huevos_ave': sum(item['total_produccion'] for item in datos['aves']) / len(datos['aves']) if datos['aves'] else 0,
        }
        
        return datos

    def obtener_datos_inventario(self):
        """
        Obtiene datos de inventario de lotes
        """
        lotes = LoteAves.objects.filter(is_active=True)
        datos = {}
        
        datos['lotes'] = list(lotes.values(
            'codigo', 'nombre_lote', 'raza', 'linea', 'numero_aves_actual', 
            'estado', 'fecha_ingreso'
        ))
        
        datos['resumen'] = {
            'total_lotes': lotes.count(),
            'total_aves': lotes.aggregate(total=Sum('numero_aves_actual'))['total'] or 0,
        }
        
        return datos

    def generar_excel(self, datos, nombre_archivo):
        """
        Genera un reporte en formato Excel.
        """
        if not OPENPYXL_AVAILABLE:
            raise ImportError("openpyxl no está instalado. Instala con: pip install openpyxl")
        
        wb = openpyxl.Workbook()
        
        # Crear hojas según tipo de reporte
        if self.tipo_reporte == 'produccion':
            self._crear_hoja_produccion_excel(wb, datos)
        elif self.tipo_reporte == 'inventario':
            self._crear_hoja_inventario_excel(wb, datos)
        
        # Guardar en buffer
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        # Crear respuesta HTTP
        response = HttpResponse(
            buffer,
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{nombre_archivo}.xlsx"'
        
        return response
    
    def _crear_hoja_produccion_excel(self, wb, datos):
        """Crea hoja de producción en Excel."""
        ws = wb.active
        ws.title = "Producción"
        
        # Encabezados
        encabezados = ['Lote', 'Fecha', 'Total Producción', 'AAA', 'AA', 'A', 'B', 'C']
        for col, encabezado in enumerate(encabezados, 1):
            cell = ws.cell(row=1, column=col, value=encabezado)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        
        # Datos
        if 'lotes' in datos:
            for row, item in enumerate(datos['lotes'], 2):
                ws.cell(row=row, column=1, value=item.get('lote__codigo', ''))
                ws.cell(row=row, column=2, value=str(item.get('fecha', '')))
                ws.cell(row=row, column=3, value=item.get('total_produccion', 0))
                ws.cell(row=row, column=4, value=item.get('produccion_aaa', 0))
                ws.cell(row=row, column=5, value=item.get('produccion_aa', 0))
                ws.cell(row=row, column=6, value=item.get('produccion_a', 0))
                ws.cell(row=row, column=7, value=item.get('produccion_b', 0))
                ws.cell(row=row, column=8, value=item.get('produccion_c', 0))

    def generar_csv(self, datos, nombre_archivo):
        """
        Genera un reporte en formato CSV.
        """
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = f'attachment; filename="{nombre_archivo}.csv"'
        
        writer = csv.writer(response)
        
        # Escribir según tipo de reporte
        if self.tipo_reporte == 'produccion' and 'lotes' in datos:
            # Encabezados
            writer.writerow(['Lote', 'Fecha', 'Total Producción', 'AAA', 'AA', 'A', 'B', 'C'])
            
            # Datos
            for item in datos['lotes']:
                writer.writerow([
                    item.get('lote__codigo', ''),
                    item.get('fecha', ''),
                    item.get('total_produccion', 0),
                    item.get('produccion_aaa', 0),
                    item.get('produccion_aa', 0),
                    item.get('produccion_a', 0),
                    item.get('produccion_b', 0),
                    item.get('produccion_c', 0),
                ])
        
        return response


class ReportePersonalizado:
    """
    Clase para crear reportes personalizados con filtros dinámicos.
    """
    
    def __init__(self, usuario):
        self.usuario = usuario
        self.filtros = []
        self.campos_seleccionados = []
        self.agrupacion = []
        self.ordenamiento = []
    
    def agregar_filtro(self, campo, operador, valor):
        """
        Agrega un filtro al reporte.
        
        Args:
            campo: Campo del modelo a filtrar
            operador: Operador de comparación ('eq', 'gt', 'lt', 'contains', etc.)
            valor: Valor a comparar
        """
        self.filtros.append({
            'campo': campo,
            'operador': operador,
            'valor': valor
        })
    
    def seleccionar_campos(self, campos):
        """Selecciona los campos a incluir en el reporte."""
        self.campos_seleccionados = campos
    
    def agrupar_por(self, campos):
        """Define campos para agrupar los resultados."""
        self.agrupacion = campos
    
    def ordenar_por(self, campos):
        """Define el orden de los resultados."""
        self.ordenamiento = campos
    
    def generar_consulta(self, modelo):
        """
        Genera la consulta Django basada en los filtros y parámetros.
        """
        queryset = modelo.objects.all()
        
        # Aplicar filtros
        for filtro in self.filtros:
            campo = filtro['campo']
            operador = filtro['operador']
            valor = filtro['valor']
            
            if operador == 'eq':
                queryset = queryset.filter(**{campo: valor})
            elif operador == 'gt':
                queryset = queryset.filter(**{f"{campo}__gt": valor})
            elif operador == 'lt':
                queryset = queryset.filter(**{f"{campo}__lt": valor})
            elif operador == 'contains':
                queryset = queryset.filter(**{f"{campo}__icontains": valor})
            elif operador == 'range':
                if isinstance(valor, (list, tuple)) and len(valor) == 2:
                    queryset = queryset.filter(**{f"{campo}__range": valor})
        
        # Aplicar selección de campos
        if self.campos_seleccionados:
            queryset = queryset.values(*self.campos_seleccionados)
        
        # Aplicar ordenamiento
        if self.ordenamiento:
            queryset = queryset.order_by(*self.ordenamiento)
        
        return queryset


def generar_reporte_automatico(reporte_programado):
    """
    Genera un reporte automático basado en la configuración programada.
    
    Args:
        reporte_programado: Instancia del modelo ReporteProgramado
    
    Returns:
        dict: Resultado de la generación del reporte
    """
    try:
        # Obtener parámetros del reporte
        parametros = {
            'fecha_inicio': reporte_programado.fecha_inicio,
            'fecha_fin': reporte_programado.fecha_fin,
        }
        
        # Crear generador de reportes
        generador = GeneradorReportes(
            tipo_reporte=reporte_programado.tipo_reporte,
            parametros=parametros
        )
        
        # Obtener datos según el tipo
        if reporte_programado.tipo_reporte == 'produccion':
            datos = generador.obtener_datos_produccion()
        elif reporte_programado.tipo_reporte == 'inventario':
            datos = generador.obtener_datos_inventario()
        else:
            raise ValueError(f"Tipo de reporte no soportado: {reporte_programado.tipo_reporte}")
        
        # Generar archivo según formato
        nombre_archivo = f"{reporte_programado.tipo_reporte}_{timezone.now().strftime('%Y%m%d_%H%M%S')}"
        
        if reporte_programado.formato == 'excel':
            archivo = generador.generar_excel(datos, nombre_archivo)
        elif reporte_programado.formato == 'csv':
            archivo = generador.generar_csv(datos, nombre_archivo)
        else:
            raise ValueError(f"Formato no soportado: {reporte_programado.formato}")
        
        # Guardar archivo si es necesario
        if reporte_programado.guardar_archivo:
            ruta_archivo = os.path.join(
                settings.MEDIA_ROOT, 
                'reportes', 
                f"{nombre_archivo}.{reporte_programado.formato}"
            )
            
            # Crear directorio si no existe
            os.makedirs(os.path.dirname(ruta_archivo), exist_ok=True)
            
            # Guardar archivo
            with open(ruta_archivo, 'wb') as f:
                f.write(archivo.content)
        
        # Enviar por email si está configurado
        if reporte_programado.enviar_email and reporte_programado.emails_destino:
            emails = [email.strip() for email in reporte_programado.emails_destino.split(',')]
            enviar_reporte_por_email(archivo, emails)
        
        return {
            'exito': True,
            'mensaje': 'Reporte generado exitosamente',
            'archivo': archivo,
            'datos': datos
        }
        
    except Exception as e:
        return {
            'exito': False,
            'mensaje': f'Error generando reporte: {str(e)}',
            'error': str(e)
        }


def enviar_reporte_por_email(reporte_generado, emails_destino):
    """
    Envía un reporte por email.
    
    Args:
        reporte_generado: Archivo del reporte generado
        emails_destino: Lista de emails destino
    """
    try:
        from django.core.mail import EmailMessage
        from django.conf import settings
        
        # Crear mensaje de email
        asunto = f"Reporte AgroSmart - {timezone.now().strftime('%d/%m/%Y')}"
        mensaje = """
        Estimado usuario,
        
        Se adjunta el reporte solicitado generado automáticamente por el sistema AgroSmart.
        
        Fecha de generación: {}
        
        Saludos,
        Sistema AgroSmart
        """.format(timezone.now().strftime('%d/%m/%Y %H:%M'))
        
        # Crear email
        email = EmailMessage(
            subject=asunto,
            body=mensaje,
            from_email=settings.DEFAULT_FROM_EMAIL,
            to=emails_destino
        )

        
        # Enviar
        email.send()
        
        return True
        
    except Exception as e:
        print(f"Error enviando email: {e}")
        return False