"""
Vistas para el sistema de reportes del módulo avícola
"""

from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.http import JsonResponse, HttpResponse
from django.views.decorators.http import require_http_methods
from django.db.models import Q, Sum, Avg, Count, F
from django.utils import timezone
from datetime import datetime, timedelta
import json
import csv
import io

from apps.usuarios.decorators import acceso_modulo_aves_required
from .models import LoteAves, BitacoraDiaria, MovimientoHuevos, ControlConcentrado, PlanVacunacion, AlertaSistema
from .reports import ReporteAvicola, ReporteComparativo, obtener_datos_dashboard

# Importaciones para Excel
try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

@login_required
@acceso_modulo_aves_required
def dashboard_reportes(request):
    """
    Dashboard principal de reportes avícolas
    """
    try:
        datos_dashboard = obtener_datos_dashboard()
    except Exception as e:
        datos_dashboard = {
            'total_lotes_activos': 0,
            'total_aves': 0,
            'huevos_hoy': 0,
            'mortalidad_hoy': 0,
            'total_huevos_30_dias': 0,
            'promedio_diario_30_dias': 0,
            'alertas_activas': 0,
            'evolucion_produccion': [],
            'top_lotes': []
        }
    
    # Obtener lotes activos para filtros
    lotes_activos = LoteAves.objects.exclude(estado='finalizado')
    
    context = {
        'datos_dashboard': datos_dashboard,
        'lotes_activos': lotes_activos,
    }
    
    return render(request, 'aves/reportes_dashboard.html', context)

@login_required
@acceso_modulo_aves_required
def generar_reporte_produccion(request):
    """
    Genera reporte de producción con filtros avanzados
    """
    # Obtener parámetros de filtro
    lote_id = request.GET.get('lote_id')
    fecha_inicio = request.GET.get('fecha_inicio')
    fecha_fin = request.GET.get('fecha_fin')
    formato = request.GET.get('formato', 'html')
    
    # Crear instancia del reporte con parámetros
    parametros = {
        'lote_id': lote_id,
        'fecha_inicio': fecha_inicio,
        'fecha_fin': fecha_fin
    }
    reporte = ReporteAvicola(parametros)
    
    # Generar datos del reporte
    try:
        datos_produccion = reporte.obtener_datos_produccion_diaria()
        resumen = reporte.obtener_resumen_produccion()
        
        # Convertir QuerySet a lista de diccionarios para el template
        datos_lista = []
        for bitacora in datos_produccion:
            datos_lista.append({
                'fecha': bitacora.fecha,
                'lote': bitacora.lote.codigo,
                'galpon': bitacora.lote.galpon,
                'produccion_aaa': bitacora.produccion_aaa,
                'produccion_aa': bitacora.produccion_aa,
                'produccion_a': bitacora.produccion_a,
                'produccion_b': bitacora.produccion_b,
                'produccion_c': bitacora.produccion_c,
                'total_huevos': bitacora.produccion_total,
                'mortalidad': bitacora.mortalidad,
                'porcentaje_postura': bitacora.porcentaje_postura,
                'consumo_concentrado': bitacora.consumo_concentrado,
                'observaciones': bitacora.observaciones
            })
        
        datos = {
            'datos_diarios': datos_lista,
            'resumen': resumen,
            'filtros': parametros
        }
    except Exception as e:
        datos = {
            'datos_diarios': [],
            'resumen': {},
            'filtros': parametros,
            'error': str(e)
        }
    
    # Exportar según formato solicitado
    if formato == 'pdf':
        return reporte.generar_pdf_produccion()
    elif formato == 'excel':
        return reporte.generar_excel_produccion()
    elif formato == 'csv':
        return reporte.generar_csv_produccion()
    
    # Renderizar HTML por defecto
    context = {
        'datos': datos,
        'lotes': LoteAves.objects.exclude(estado='finalizado'),
        'filtros': parametros
    }
    
    return render(request, 'aves/reporte_produccion.html', context)

@login_required
@acceso_modulo_aves_required
def reporte_comparativo_lotes(request):
    """
    Vista para el reporte comparativo entre lotes
    """
    # Obtener parámetros
    lotes_ids = request.GET.getlist('lotes')
    fecha_inicio = request.GET.get('fecha_inicio')
    fecha_fin = request.GET.get('fecha_fin')
    formato = request.GET.get('formato', 'html')
    
    try:
        reporte_comparativo = ReporteComparativo()
        datos_comparacion = reporte_comparativo.comparar_lotes(lotes_ids, fecha_inicio, fecha_fin)
        
        # Exportar según formato
        if formato == 'excel':
            return generar_excel_comparativo(datos_comparacion)
    except Exception as e:
        datos_comparacion = []
    
    context = {
        'datos_comparacion': datos_comparacion,
        'lotes': LoteAves.objects.exclude(estado='finalizado'),
        'filtros': {
            'lotes_ids': lotes_ids,
            'fecha_inicio': fecha_inicio,
            'fecha_fin': fecha_fin,
        }
    }
    
    return render(request, 'aves/reporte_comparativo_lotes.html', context)

@login_required
@acceso_modulo_aves_required
def reporte_mortalidad(request):
    """
    Vista para generar reportes de mortalidad
    """
    context = {
        'titulo': 'Reporte de Mortalidad',
        'lotes_activos': LoteAves.objects.exclude(estado='finalizado')
    }
    
    # Obtener parámetros de filtro
    lote_id = request.GET.get('lote_id')
    fecha_inicio = request.GET.get('fecha_inicio')
    fecha_fin = request.GET.get('fecha_fin')
    formato = request.GET.get('formato', 'html')
    
    if lote_id or fecha_inicio or fecha_fin:
        try:
            # Construir queryset con filtros
            queryset = BitacoraDiaria.objects.select_related('lote')
            
            if lote_id:
                queryset = queryset.filter(lote_id=lote_id)
            if fecha_inicio:
                queryset = queryset.filter(fecha__gte=fecha_inicio)
            if fecha_fin:
                queryset = queryset.filter(fecha__lte=fecha_fin)
            
            # Obtener datos de mortalidad
            datos_mortalidad = []
            total_mortalidad = 0
            
            for bitacora in queryset.order_by('-fecha'):
                datos_mortalidad.append({
                    'fecha': bitacora.fecha,
                    'lote': bitacora.lote.codigo,
                    'galpon': bitacora.lote.galpon,
                    'mortalidad': bitacora.mortalidad,
                    'causa_mortalidad': bitacora.causa_mortalidad,
                    'aves_actuales': bitacora.lote.numero_aves_actual,
                    'porcentaje_mortalidad': round((bitacora.mortalidad / bitacora.lote.numero_aves_actual * 100), 2) if bitacora.lote.numero_aves_actual > 0 else 0
                })
                total_mortalidad += bitacora.mortalidad
            
            # Calcular resumen
            resumen = {
                'total_mortalidad': total_mortalidad,
                'registros_encontrados': len(datos_mortalidad),
                'promedio_diario': round(total_mortalidad / len(datos_mortalidad), 2) if datos_mortalidad else 0
            }
            
            # Exportar según formato
            if formato == 'excel':
                return generar_excel_mortalidad(datos_mortalidad, resumen)
            elif formato == 'csv':
                return generar_csv_mortalidad(datos_mortalidad)
            
            context.update({
                'datos_mortalidad': datos_mortalidad,
                'resumen': resumen,
                'filtros': {
                    'lote_id': lote_id,
                    'fecha_inicio': fecha_inicio,
                    'fecha_fin': fecha_fin
                }
            })
                
        except Exception as e:
            context['error'] = f'Error al generar el reporte: {str(e)}'
    
    return render(request, 'aves/reporte_mortalidad.html', context)

@login_required
@acceso_modulo_aves_required
@require_http_methods(["GET"])
def api_datos_dashboard(request):
    """
    API para obtener datos del dashboard en tiempo real
    """
    try:
        datos = obtener_datos_dashboard()
        return JsonResponse(datos)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)

@login_required


@login_required
@acceso_modulo_aves_required
def exportar_datos_completos(request):
    """
    Exporta todos los datos del sistema en formato Excel o CSV
    """
    formato = request.GET.get('formato', 'excel')
    incluir_historicos = request.GET.get('incluir_historicos', 'false') == 'true'
    
    if formato == 'excel':
        return generar_excel_datos_completos(incluir_historicos)
    elif formato == 'csv':
        return generar_csv_datos_completos(incluir_historicos)
    else:
        return JsonResponse({'error': 'Formato no válido'}, status=400)

@login_required
@acceso_modulo_aves_required
def reporte_salud_vacunacion(request):
    """
    Vista para el reporte de salud y vacunación
    """
    # Obtener parámetros de filtro
    lote_id = request.GET.get('lote_id')
    fecha_inicio = request.GET.get('fecha_inicio')
    fecha_fin = request.GET.get('fecha_fin')
    formato = request.GET.get('formato', 'html')
    
    context = {
        'titulo': 'Reporte de Salud y Vacunación',
        'lotes_activos': LoteAves.objects.exclude(estado='finalizado')
    }
    
    if lote_id or fecha_inicio or fecha_fin:
        try:
            # Construir queryset con filtros
            queryset = PlanVacunacion.objects.select_related('lote', 'tipo_vacuna', 'veterinario')
            
            if lote_id:
                queryset = queryset.filter(lote_id=lote_id)
            if fecha_inicio:
                queryset = queryset.filter(fecha_programada__gte=fecha_inicio)
            if fecha_fin:
                queryset = queryset.filter(fecha_programada__lte=fecha_fin)
            
            # Obtener datos de vacunación
            datos_vacunacion = []
            aplicadas = 0
            pendientes = 0
            
            for plan in queryset.order_by('-fecha_programada'):
                datos_vacunacion.append({
                    'fecha_programada': plan.fecha_programada,
                    'fecha_aplicada': plan.fecha_aplicada,
                    'lote': plan.lote.codigo,
                    'galpon': plan.lote.galpon,
                    'tipo_vacuna': plan.tipo_vacuna.nombre,
                    'veterinario': plan.veterinario.get_full_name() or plan.veterinario.username,
                    'aves_vacunadas': plan.numero_aves_vacunadas,
                    'aplicada': plan.aplicada,
                    'observaciones': plan.observaciones
                })
                
                if plan.aplicada:
                    aplicadas += 1
                else:
                    pendientes += 1
            
            # Calcular resumen
            resumen = {
                'total_vacunas': len(datos_vacunacion),
                'aplicadas': aplicadas,
                'pendientes': pendientes,
                'porcentaje_cumplimiento': round((aplicadas / len(datos_vacunacion) * 100), 2) if datos_vacunacion else 0
            }
            
            # Exportar según formato
            if formato == 'excel':
                return generar_excel_vacunacion(datos_vacunacion, resumen)
            elif formato == 'csv':
                return generar_csv_vacunacion(datos_vacunacion)
            
            context.update({
                'datos_vacunacion': datos_vacunacion,
                'resumen': resumen,
                'filtros': {
                    'lote_id': lote_id,
                    'fecha_inicio': fecha_inicio,
                    'fecha_fin': fecha_fin
                }
            })
                
        except Exception as e:
            context['error'] = f'Error al generar el reporte: {str(e)}'
    
    return render(request, 'aves/reporte_vacunacion.html', context)

@login_required
@acceso_modulo_aves_required
def reporte_consumo_concentrado(request):
    """
    Vista para el reporte de consumo de concentrado
    """
    # Obtener parámetros de filtro
    lote_id = request.GET.get('lote_id')
    fecha_inicio = request.GET.get('fecha_inicio')
    fecha_fin = request.GET.get('fecha_fin')
    formato = request.GET.get('formato', 'html')
    
    context = {
        'titulo': 'Reporte de Consumo de Concentrado',
        'lotes_activos': LoteAves.objects.exclude(estado='finalizado')
    }
    
    if lote_id or fecha_inicio or fecha_fin:
        try:
            # Construir queryset con filtros
            queryset = BitacoraDiaria.objects.select_related('lote')
            
            if lote_id:
                queryset = queryset.filter(lote_id=lote_id)
            if fecha_inicio:
                queryset = queryset.filter(fecha__gte=fecha_inicio)
            if fecha_fin:
                queryset = queryset.filter(fecha__lte=fecha_fin)
            
            # Obtener datos de consumo
            datos_consumo = []
            total_consumo = 0
            
            for bitacora in queryset.order_by('-fecha'):
                consumo_por_ave = round(bitacora.consumo_concentrado / bitacora.lote.numero_aves_actual, 3) if bitacora.lote.numero_aves_actual > 0 else 0
                datos_consumo.append({
                    'fecha': bitacora.fecha,
                    'lote': bitacora.lote.codigo,
                    'galpon': bitacora.lote.galpon,
                    'consumo_total': bitacora.consumo_concentrado,
                    'aves_actuales': bitacora.lote.numero_aves_actual,
                    'consumo_por_ave': consumo_por_ave
                })
                total_consumo += bitacora.consumo_concentrado
            
            # Calcular resumen
            resumen = {
                'total_consumo': total_consumo,
                'registros_encontrados': len(datos_consumo),
                'promedio_diario': round(total_consumo / len(datos_consumo), 2) if datos_consumo else 0
            }
            
            # Exportar según formato
            if formato == 'excel':
                return generar_excel_consumo(datos_consumo, resumen)
            elif formato == 'csv':
                return generar_csv_consumo(datos_consumo)
            
            context.update({
                'datos_consumo': datos_consumo,
                'resumen': resumen,
                'filtros': {
                    'lote_id': lote_id,
                    'fecha_inicio': fecha_inicio,
                    'fecha_fin': fecha_fin
                }
            })
                
        except Exception as e:
            context['error'] = f'Error al generar el reporte: {str(e)}'
    
    return render(request, 'aves/reporte_consumo.html', context)





def generar_excel_datos_completos(incluir_historicos=False):
    """Genera Excel con todos los datos del sistema"""
    if not OPENPYXL_AVAILABLE:
        return HttpResponse("OpenPyXL no está disponible", status=500)
    
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="datos_completos.xlsx"'
    
    wb = openpyxl.Workbook()
    
    # Hoja de Lotes
    ws_lotes = wb.active
    ws_lotes.title = "Lotes"
    
    headers_lotes = ['Código', 'Galpón', 'Línea Genética', 'Aves Actuales', 'Estado', 'Fecha Llegada']
    for col, header in enumerate(headers_lotes, 1):
        ws_lotes.cell(row=1, column=col, value=header)
    
    lotes = LoteAves.objects.all()
    for row, lote in enumerate(lotes, 2):
        ws_lotes.cell(row=row, column=1, value=lote.codigo)
        ws_lotes.cell(row=row, column=2, value=lote.galpon)
        ws_lotes.cell(row=row, column=3, value=lote.get_linea_genetica_display())
        ws_lotes.cell(row=row, column=4, value=lote.numero_aves_actual)
        ws_lotes.cell(row=row, column=5, value=lote.get_estado_display())
        ws_lotes.cell(row=row, column=6, value=lote.fecha_llegada.strftime('%d/%m/%Y'))
    
    wb.save(response)
    return response

def generar_csv_datos_completos(incluir_historicos=False):
    """Genera CSV con todos los datos del sistema"""
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="datos_completos.csv"'
    
    writer = csv.writer(response)
    
    # Escribir datos de lotes
    writer.writerow(['=== LOTES ==='])
    writer.writerow(['Código', 'Galpón', 'Línea Genética', 'Aves Actuales', 'Estado', 'Fecha Llegada'])
    
    lotes = LoteAves.objects.all()
    for lote in lotes:
        writer.writerow([
            lote.codigo,
            lote.galpon,
            lote.get_linea_genetica_display(),
            lote.numero_aves_actual,
            lote.get_estado_display(),
            lote.fecha_llegada.strftime('%d/%m/%Y')
        ])
    
    return response


@login_required
@acceso_modulo_aves_required
def generar_reporte_sena(request):
    """
    Genera reporte mensual en formato SENA
    """
    if request.method == 'GET':
        # Mostrar formulario de selección
        lotes = LoteAves.objects.filter(is_active=True)
        
        # Obtener mes y año actual por defecto
        hoy = timezone.now().date()
        mes_actual = hoy.month
        año_actual = hoy.year
        
        context = {
            'lotes': lotes,
            'mes_actual': mes_actual,
            'año_actual': año_actual,
            'meses': [
                (1, 'Enero'), (2, 'Febrero'), (3, 'Marzo'), (4, 'Abril'),
                (5, 'Mayo'), (6, 'Junio'), (7, 'Julio'), (8, 'Agosto'),
                (9, 'Septiembre'), (10, 'Octubre'), (11, 'Noviembre'), (12, 'Diciembre')
            ],
            'años': list(range(año_actual - 2, año_actual + 1))
        }
        
        return render(request, 'aves/reporte_sena_form.html', context)
    
    elif request.method == 'POST':
        # Generar reporte
        lote_id = request.POST.get('lote_id')
        mes = int(request.POST.get('mes'))
        año = int(request.POST.get('año'))
        nombre_granja = request.POST.get('nombre_granja', 'Granja Avícola La Salada')
        registro_ica = request.POST.get('registro_ica', '051290274')
        
        try:
            from .reports import generar_reporte_sena_excel
            return generar_reporte_sena_excel(lote_id, mes, año, nombre_granja, registro_ica)
        except Exception as e:
            messages.error(request, f'Error al generar el reporte: {str(e)}')
            return redirect('aves:generar_reporte_sena')