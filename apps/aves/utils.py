"""
Utilidades para el módulo avícola.
"""

from django.http import HttpResponse, HttpResponseServerError
import traceback
from calendar import monthrange

from .models import AlertaSistema, InventarioHuevos


def generar_alertas(bitacora_instance=None):
    """Genera alertas automáticas del sistema."""
    alertas_generadas = []
    
    if bitacora_instance:
        # Alerta por baja producción basada en la bitácora
        lote = bitacora_instance.lote
        total_produccion = (
            bitacora_instance.produccion_aaa + 
            bitacora_instance.produccion_aa + 
            bitacora_instance.produccion_a + 
            bitacora_instance.produccion_b + 
            bitacora_instance.produccion_c
        )
        
        # Ejemplo: alerta si la producción es muy baja
        if total_produccion < 100:  # Ajustar según criterios
            try:
                alerta, created = AlertaSistema.objects.get_or_create(
                    tipo='baja_produccion',
                    lote=lote,
                    fecha=bitacora_instance.fecha,
                    defaults={
                        'mensaje': f'Baja producción detectada en lote {lote.codigo}: {total_produccion} huevos',
                        'nivel': 'warning',
                        'activa': True
                    }
                )
                if created:
                    alertas_generadas.append(alerta)
            except Exception:
                pass
        
        # Alerta por alta mortalidad
        if bitacora_instance.mortalidad > 5:  # Ajustar según criterios
            try:
                alerta, created = AlertaSistema.objects.get_or_create(
                    tipo='alta_mortalidad',
                    lote=lote,
                    fecha=bitacora_instance.fecha,
                    defaults={
                        'mensaje': f'Alta mortalidad detectada en lote {lote.codigo}: {bitacora_instance.mortalidad} aves',
                        'nivel': 'danger',
                        'activa': True
                    }
                )
                if created:
                    alertas_generadas.append(alerta)
            except Exception:
                pass
    
    return alertas_generadas


def actualizar_inventario_huevos(bitacora_instance):
    """Actualiza el inventario de huevos automáticamente."""
    try:
        # Mapeo de categorías de huevos
        categorias_produccion = {
            'AAA': bitacora_instance.produccion_aaa,
            'AA': bitacora_instance.produccion_aa,
            'A': bitacora_instance.produccion_a,
            'B': bitacora_instance.produccion_b,
            'C': bitacora_instance.produccion_c,
        }
        
        # Actualizar inventario por cada categoría
        for categoria, cantidad in categorias_produccion.items():
            if cantidad > 0:  # Solo actualizar si hay producción
                inventario, created = InventarioHuevos.objects.get_or_create(
                    categoria=categoria,
                    defaults={
                        'cantidad_actual': 0,
                        'cantidad_minima': 100,
                        'stock_automatico': True,
                        'factor_calculo': 0.75,
                        'dias_stock': 3,
                    }
                )
                
                # Incrementar la cantidad actual
                inventario.cantidad_actual += cantidad
                inventario.save()
        
        return True
    except Exception as e:
        print(f"Error actualizando inventario: {e}")
        return False


def actualizar_inventario_por_movimiento(detalle_movimiento):
    """Actualiza el inventario cuando se registra un movimiento de huevos."""
    try:
        # Obtener o crear el inventario para la categoría
        inventario, created = InventarioHuevos.objects.get_or_create(
            categoria=detalle_movimiento.categoria_huevo,
            defaults={
                'cantidad_actual': 0,
                'cantidad_minima': 100,
                'stock_automatico': True,
                'factor_calculo': 0.75,
                'dias_stock': 3,
            }
        )
        
        # Determinar si es una salida o entrada basado en el tipo de movimiento
        if hasattr(detalle_movimiento, 'movimiento'):
            tipo_movimiento = detalle_movimiento.movimiento.tipo_movimiento
        else:
            # Para casos de reversión (cuando se elimina un movimiento)
            tipo_movimiento = 'devolucion'  # Asumimos que es una devolución
        
        # Restar la cantidad movida (en unidades) solo para salidas
        cantidad_unidades = detalle_movimiento.cantidad_unidades
        
        if tipo_movimiento in ['venta', 'autoconsumo', 'baja']:
            # Es una salida - restar del inventario
            inventario.cantidad_actual -= cantidad_unidades
            print(f"Salida registrada: -{cantidad_unidades} unidades de {detalle_movimiento.categoria_huevo}")
        else:  # devolución
            # Es una entrada - sumar al inventario
            inventario.cantidad_actual += cantidad_unidades
            print(f"Entrada registrada: +{cantidad_unidades} unidades de {detalle_movimiento.categoria_huevo}")
        
        # Verificar que no quede en negativo para salidas
        if inventario.cantidad_actual < 0 and tipo_movimiento in ['venta', 'autoconsumo', 'baja']:
            print(f"Advertencia: Stock insuficiente para {detalle_movimiento.categoria_huevo}. "
                  f"Stock resultante: {inventario.cantidad_actual}")
            # Permitir que quede en negativo para registrar el faltante
        
        inventario.save()
        return True
        
    except Exception as e:
        print(f"Error actualizando inventario por movimiento: {e}")
        return False

def exportar_reporte_excel(tipo_reporte, datos, estadisticas, filtros=None):
    """Exporta reportes a Excel en formato SENA oficial exacto."""
    try:
        # Import perezoso: solo cuando realmente se exporta
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from openpyxl.drawing import image
        import os
        from django.conf import settings

        # Crear workbook y worksheet
        wb = openpyxl.Workbook()
        ws = wb.active
        
        # Obtener información del primer registro para determinar mes/año
        if datos and len(datos) > 0:
            primera_fecha = datos[0].fecha
            mes = primera_fecha.month
            año = primera_fecha.year
            lote = datos[0].lote
        else:
            # Valores por defecto si no hay datos
            from datetime import datetime
            hoy = datetime.now()
            mes = hoy.month
            año = hoy.year
            lote = None
        
        ws.title = f"Registro Postura {mes}-{año}"
        
        # Configurar página
        ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
        ws.page_setup.paperSize = ws.PAPERSIZE_A4
        
        # Estilos exactos del formato SENA
        header_font = Font(name='Arial', size=10, bold=True)
        normal_font = Font(name='Arial', size=9)
        title_font = Font(name='Arial', size=11, bold=True)
        small_font = Font(name='Arial', size=8)
        
        border_thin = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        border_thick = Border(
            left=Side(style='thick'),
            right=Side(style='thick'),
            top=Side(style='thick'),
            bottom=Side(style='thick')
        )
        
        # Colores exactos del formato SENA
        sena_green = PatternFill(start_color='30A900', end_color='30A900', fill_type='solid')
        light_gray = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
        
        # LOGO SENA - Insertar imagen si existe
        try:
            logo_path = os.path.join(settings.STATIC_ROOT or settings.STATICFILES_DIRS[0], 'img', 'logo-ico.ico')
            if not os.path.exists(logo_path):
                logo_path = os.path.join(settings.BASE_DIR, 'static', 'img', 'logo-ico.ico')
            
            if os.path.exists(logo_path):
                img = image.Image(logo_path)
                img.width = 100
                img.height = 100
                ws.add_image(img, 'A1')
        except:
            # Si no se puede cargar la imagen, usar texto
            pass
                # Función helper para aplicar bordes a celdas combinadas
        def aplicar_borde_completo(ws, rango, border_style):
            """Aplica bordes a todas las celdas de un rango, incluso si están combinadas"""
            from openpyxl.utils import range_boundaries
            min_col, min_row, max_col, max_row = range_boundaries(rango)
            
            for row in range(min_row, max_row + 1):
                for col in range(min_col, max_col + 1):
                    ws.cell(row=row, column=col).border = border_style

        # ENCABEZADO PRINCIPAL - Fila 1
        ws.merge_cells('A1:A3')  # Espacio para logo SENA
        ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
        aplicar_borde_completo(ws, 'A1:A3', border_thick)
        

        # Luego usar así:
        ws.merge_cells('B1:M3')
        ws['B1'] = 'Granja Avícola La Salada'
        ws['B1'].font = Font(name='Arial', size=14, bold=True)
        ws['B1'].alignment = Alignment(horizontal='center', vertical='center')
        aplicar_borde_completo(ws, 'B1:M3', border_thick)
        
        ws.merge_cells('N1:P3')
        ws['N1'] = 'CENTRO DE LOS\nRECURSOS\nNATURALES\nRENOVABLES'
        ws['N1'].font = Font(name='Arial', size=9, bold=True)
        ws['N1'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        aplicar_borde_completo(ws, 'N1:P3', border_thick)



        
        # Fila 4: Registro ICA
        ws.merge_cells('A4:P4')
        ws['A4'] = 'Registro ICA 051290274'
        ws['A4'].font = Font(name='Arial', size=11, bold=True)
        ws['A4'].alignment = Alignment(horizontal='center', vertical='center')
        aplicar_borde_completo(ws, 'A4:P4', border_thick)
        
        # Fila 5: Vacía
        ws.row_dimensions[5].height = 5
        
        # Fila 6: Título del reporte
        ws.merge_cells('A6:P6')
        ws['A6'] = 'REGISTRO MENSUAL DE POSTURA'
        ws['A6'].font = Font(name='Arial', size=12, bold=True)
        ws['A6'].alignment = Alignment(horizontal='center', vertical='center')
        aplicar_borde_completo(ws, 'A6:P6', border_thick)
        ws['A6'].fill = sena_green
        
        # Fila 7: Versión
        ws.merge_cells('A7:P7')
        ws['A7'] = 'Versión: 2020-01'
        ws['A7'].font = Font(name='Arial', size=8)
        ws['A7'].alignment = Alignment(horizontal='center', vertical='center')
        aplicar_borde_completo(ws, 'A7:P7', border_thick)
        
        # Fila 8: Información del lote - Primera línea
        info_labels = [
            ('A8', 'Mes:', 'B8'),
            ('C8', 'Galpón:', 'D8'), 
            ('E8', 'Sistema:', 'F8'),
            ('G8', 'Línea:', 'H8')
        ]
        
        for label_cell, label_text, value_cell in info_labels:
            ws[label_cell] = label_text
            ws[label_cell].font = header_font
            ws[label_cell].border = border_thin
            ws[label_cell].fill = light_gray
            
            if label_text == 'Mes:':
                ws[value_cell] = f'{mes:02d}'
            elif label_text == 'Galpón:':
                ws[value_cell] = str(lote.galpon) if lote else ''
            elif label_text == 'Sistema:':
                ws[value_cell] = 'Jaula'
            elif label_text == 'Línea:':
                ws[value_cell] = str(lote.linea_genetica) if lote and hasattr(lote, 'linea_genetica') else ''
            
            ws[value_cell].font = normal_font
            aplicar_borde_completo(ws, value_cell, border_thin)
        
        # Fila 9: Información del lote - Segunda línea
        ws['A9'] = 'Edad en semanas:'
        ws['A9'].font = header_font
        aplicar_borde_completo(ws, 'A9', border_thin)
        ws['A9'].fill = light_gray
        
        edad_semanas = getattr(lote, 'edad_actual_semanas', 0) if lote else 0
        ws['B9'] = f'{edad_semanas:.1f}'
        ws['B9'].font = normal_font
        aplicar_borde_completo(ws, 'B9', border_thin)
        
        ws['C9'] = 'Aves alojadas:'
        ws['C9'].font = header_font
        aplicar_borde_completo(ws, 'C9', border_thin)
        ws['C9'].fill = light_gray
        
        ws['D9'] = lote.numero_aves_inicial if lote else 0
        ws['D9'].font = normal_font
        aplicar_borde_completo(ws, 'D9', border_thin)
        
        ws['E9'] = 'N° de aves al inicio del mes:'
        ws['E9'].font = header_font
        aplicar_borde_completo(ws, 'E9', border_thin)
        ws['E9'].fill = light_gray
        
        ws['F9'] = lote.numero_aves_actual if lote else 0
        ws['F9'].font = normal_font
        aplicar_borde_completo(ws, 'F9', border_thin)
        
        # Fila 10: Vacía
        ws.row_dimensions[10].height = 5
        
        # TABLA DE DATOS DIARIOS - Fila 11: Encabezados principales
        # Día
        ws.merge_cells('A11:A12')
        ws['A11'] = 'Día'
        ws['A11'].font = header_font
        ws['A11'].alignment = Alignment(horizontal='center', vertical='center')
        ws['A11'].fill = sena_green
        aplicar_borde_completo(ws, 'A11:A12', border_thick)
        
        # PRODUCCIÓN (expandir para incluir promedio)
        ws.merge_cells('B11:G11')
        ws['B11'] = 'PRODUCCIÓN'
        ws['B11'].font = header_font
        ws['B11'].alignment = Alignment(horizontal='center', vertical='center')
        ws['B11'].fill = sena_green
        aplicar_borde_completo(ws, 'B11:G11', border_thick)
        
        # ALIMENTO (mover a H-I)
        ws.merge_cells('H11:I11')
        ws['H11'] = 'ALIMENTO'
        ws['H11'].font = header_font
        ws['H11'].alignment = Alignment(horizontal='center', vertical='center')
        ws['H11'].fill = sena_green
        aplicar_borde_completo(ws, 'H11:I11', border_thick)
        
        # BAJAS (mover a J-L)
        ws.merge_cells('J11:L11')
        ws['J11'] = 'BAJAS'
        ws['J11'].font = header_font
        ws['J11'].alignment = Alignment(horizontal='center', vertical='center')
        ws['J11'].fill = sena_green
        aplicar_borde_completo(ws, 'J11:L11', border_thick)
        
        # Existencia (mover a M)
        ws.merge_cells('M11:M12')
        ws['M11'] = 'Existencia'
        ws['M11'].font = header_font
        ws['M11'].alignment = Alignment(horizontal='center', vertical='center')
        ws['M11'].fill = sena_green
        aplicar_borde_completo(ws, 'M11:M12', border_thick)
        
        # OBSERVACIONES (mover a N-P)
        ws.merge_cells('N11:P12')
        ws['N11'] = 'OBSERVACIONES'
        ws['N11'].font = header_font
        ws['N11'].alignment = Alignment(horizontal='center', vertical='center')
        ws['N11'].fill = sena_green
        aplicar_borde_completo(ws, 'N11:P12', border_thick)
        
        # Fila 12: Subencabezados (actualizar posiciones)
        subencabezados = [
            ('B12', '1a'),
            ('C12', '2a'),
            ('D12', '3a'),
            ('E12', 'Rotos'),
            ('F12', 'Total'),
            ('G12', 'Promedio'),  # Nueva columna de promedio
            ('H12', 'Kg'),        # Alimento movido
            ('I12', 'Acumulado'), # Alimento acumulado movido
            ('J12', 'Descar'),    # Bajas movido
            ('K12', 'Elimin'),    # Bajas movido
            ('L12', 'Total'),     # Total bajas movido
        ]
        
        for celda, texto in subencabezados:
            ws[celda] = texto
            ws[celda].font = header_font
            ws[celda].alignment = Alignment(horizontal='center', vertical='center')
            ws[celda].fill = light_gray
            aplicar_borde_completo(ws, celda, border_thin)
        
        # Obtener días del mes
        dias_en_mes = monthrange(año, mes)[1]
        
        # Crear diccionario de bitácoras por día
        bitacoras_por_dia = {}
        if datos:
            for bitacora in datos:
                if bitacora.fecha.month == mes and bitacora.fecha.year == año:
                    bitacoras_por_dia[bitacora.fecha.day] = bitacora
        
        # LLENAR DATOS DIARIOS
        fila_inicio_datos = 13
        acumulado_alimento = 0
        
        # Variables para totales
        total_1a = 0
        total_2a = 0
        total_3a = 0
        total_rotos = 0
        total_alimento = 0
        total_mortalidad = 0
        
        # Variable para calcular existencia correctamente
        aves_iniciales_mes = lote.numero_aves_actual if lote else 0
        mortalidad_acumulada = 0
        
        for dia in range(1, dias_en_mes + 1):
            fila = fila_inicio_datos + dia - 1
            
            # Día
            ws[f'A{fila}'] = dia
            ws[f'A{fila}'].alignment = Alignment(horizontal='center', vertical='center')
            ws[f'A{fila}'].border = border_thin
            ws[f'A{fila}'].font = normal_font
            
            if dia in bitacoras_por_dia:
                bitacora = bitacoras_por_dia[dia]
                
                # Producción (corregir mapeo según formato SENA)
                # 1a = AAA (primera calidad)
                produccion_1a = bitacora.produccion_aaa or 0
                # 2a = AA (segunda calidad) 
                produccion_2a = bitacora.produccion_aa or 0
                # 3a = A + B + C (tercera calidad y menores)
                produccion_3a = (bitacora.produccion_a or 0) + (bitacora.produccion_b or 0) + (bitacora.produccion_c or 0)
                rotos = bitacora.huevos_rotos or 0
                total_produccion_dia = produccion_1a + produccion_2a + produccion_3a + rotos
                
                # Acumular totales
                total_1a += produccion_1a
                total_2a += produccion_2a
                total_3a += produccion_3a
                total_rotos += rotos
                
                ws[f'B{fila}'] = produccion_1a if produccion_1a > 0 else ''
                ws[f'C{fila}'] = produccion_2a if produccion_2a > 0 else ''
                ws[f'D{fila}'] = produccion_3a if produccion_3a > 0 else ''
                ws[f'E{fila}'] = rotos if rotos > 0 else ''
                ws[f'F{fila}'] = total_produccion_dia if total_produccion_dia > 0 else ''
                
                # Calcular promedio diario de producción (porcentaje) - COLUMNA G
                promedio_dia = (total_produccion_dia / lote.numero_aves_actual * 100) if lote and lote.numero_aves_actual > 0 and total_produccion_dia > 0 else 0
                ws[f'G{fila}'] = f'{promedio_dia:.1f}%' if promedio_dia > 0 else ''
                
                # Alimento (ahora en columnas H e I)
                consumo_kg = float(bitacora.consumo_concentrado) if bitacora.consumo_concentrado else 0
                acumulado_alimento += consumo_kg
                total_alimento += consumo_kg
                
                ws[f'H{fila}'] = f'{consumo_kg:.1f}' if consumo_kg > 0 else ''
                ws[f'I{fila}'] = f'{acumulado_alimento:.1f}' if acumulado_alimento > 0 else ''
                
                # Bajas (mortalidad) - ahora en columnas J, K, L
                mortalidad_dia = bitacora.mortalidad or 0
                mortalidad_acumulada += mortalidad_dia
                total_mortalidad += mortalidad_dia
                
                ws[f'J{fila}'] = mortalidad_dia if mortalidad_dia > 0 else ''  # Descartes
                ws[f'K{fila}'] = ''  # Eliminación (no tenemos este dato específico)
                ws[f'L{fila}'] = mortalidad_dia if mortalidad_dia > 0 else ''  # Total bajas
                
                # Existencia (aves restantes) - ahora en columna M
                aves_restantes = aves_iniciales_mes - mortalidad_acumulada
                ws[f'M{fila}'] = aves_restantes if aves_restantes >= 0 else ''
                
                # Observaciones (ahora en columnas N:P)
                observaciones = bitacora.observaciones[:100] if bitacora.observaciones else ''
                ws.merge_cells(f'N{fila}:P{fila}')
                ws[f'N{fila}'] = observaciones
            else:
                # Si no hay bitácora para este día, mantener existencia del día anterior
                if dia > 1:
                    fila_anterior = fila - 1
                    try:
                        existencia_anterior = ws[f'M{fila_anterior}'].value
                        if existencia_anterior:
                            ws[f'M{fila}'] = existencia_anterior
                    except:
                        pass
            
            # Aplicar bordes y formato a toda la fila (actualizar rango de columnas)
            for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M']:
                ws[f'{col}{fila}'].border = border_thin
                ws[f'{col}{fila}'].font = normal_font
                ws[f'{col}{fila}'].alignment = Alignment(horizontal='center', vertical='center')
            
            # Formato especial para observaciones (actualizar rango)
            for col in ['N', 'O', 'P']: 
                ws[f'{col}{fila}'].border = border_thin
                ws[f'{col}{fila}'].font = normal_font
                ws[f'{col}{fila}'].alignment = Alignment(horizontal='left', vertical='center')
        
        # FILA TOTAL (actualizar todas las columnas)
        fila_total = fila_inicio_datos + dias_en_mes
        ws[f'A{fila_total}'] = 'TOTAL'
        ws[f'A{fila_total}'].font = header_font
        ws[f'A{fila_total}'].fill = sena_green
        ws[f'A{fila_total}'].alignment = Alignment(horizontal='center', vertical='center')
        aplicar_borde_completo(ws, f'A{fila_total}:P{fila_total}', border_thick)
        
        total_general = total_1a + total_2a + total_3a + total_rotos
        promedio_total = (total_general / (lote.numero_aves_actual * dias_en_mes) * 100) if lote and lote.numero_aves_actual > 0 else 0
        
        ws[f'B{fila_total}'] = total_1a if total_1a > 0 else ''
        ws[f'C{fila_total}'] = total_2a if total_2a > 0 else ''
        ws[f'D{fila_total}'] = total_3a if total_3a > 0 else ''
        ws[f'E{fila_total}'] = total_rotos if total_rotos > 0 else ''
        ws[f'F{fila_total}'] = total_general if total_general > 0 else ''
        ws[f'G{fila_total}'] = f'{promedio_total:.1f}%' if promedio_total > 0 else ''  # Promedio total
        
        # Total alimento (ahora en H)
        ws[f'H{fila_total}'] = f'{total_alimento:.1f}' if total_alimento > 0 else ''
        
        # Total bajas (ahora en L)
        ws[f'L{fila_total}'] = total_mortalidad if total_mortalidad > 0 else ''
        
        # Aplicar formato a fila total (actualizar rango de columnas)
        for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P']:
            ws[f'{col}{fila_total}'].border = border_thin
            ws[f'{col}{fila_total}'].font = header_font
            ws[f'{col}{fila_total}'].fill = sena_green
            if col in ['B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M']:
                ws[f'{col}{fila_total}'].alignment = Alignment(horizontal='center', vertical='center')
        
        # RESUMEN MENSUAL
        fila_resumen = fila_total + 2
        ws.merge_cells(f'A{fila_resumen}:P{fila_resumen}')
        ws[f'A{fila_resumen}'] = 'RESUMEN MENSUAL'
        ws[f'A{fila_resumen}'].font = Font(name='Arial', size=12, bold=True)
        ws[f'A{fila_resumen}'].fill = sena_green
        ws[f'A{fila_resumen}'].alignment = Alignment(horizontal='center', vertical='center')
        aplicar_borde_completo(ws, f'A{fila_resumen}:P{fila_resumen}', border_thick)
        
        # Calcular porcentajes y estadísticas
        if lote and lote.numero_aves_actual > 0:
            # Porcentaje de postura
            porcentaje_postura = (total_general / (lote.numero_aves_actual * dias_en_mes)) * 100
            
            # Consumo promedio por ave
            consumo_promedio = total_alimento / lote.numero_aves_actual if lote.numero_aves_actual > 0 else 0
            
            # Mortalidad porcentual
            mortalidad_porcentual = (total_mortalidad / lote.numero_aves_inicial) * 100 if lote.numero_aves_inicial > 0 else 0
            
            # Datos del resumen
            resumen_datos = [
                (f'A{fila_resumen + 1}', 'PRODUCCIÓN TOTAL:', f'B{fila_resumen + 1}', total_general, f'D{fila_resumen + 1}', '% DE PRODUCCIÓN:', f'E{fila_resumen + 1}', f'{porcentaje_postura:.1f}%'),
                (f'A{fila_resumen + 2}', 'CONSUMO TOTAL:', f'B{fila_resumen + 2}', f'{total_alimento:.1f}', f'D{fila_resumen + 2}', '% DE MORTALIDAD:', f'E{fila_resumen + 2}', f'{mortalidad_porcentual:.2f}%'),
                (f'A{fila_resumen + 3}', 'CONVERSIÓN:', f'B{fila_resumen + 3}', f'{consumo_promedio:.2f}', f'D{fila_resumen + 3}', '', f'E{fila_resumen + 3}', ''),
            ]
            
            for datos_fila in resumen_datos:
                for i in range(0, len(datos_fila), 2):
                    if i + 1 < len(datos_fila):
                        celda = datos_fila[i]
                        valor = datos_fila[i + 1]
                        ws[celda] = valor
                        ws[celda].font = header_font if 'TOTAL' in str(valor) or '%' in str(valor) else normal_font
                        aplicar_borde_completo(ws, celda, border_thin)
                        if i == 0 or i == 4:  # Labels
                            ws[celda].fill = light_gray
        
        # Fila final con firmas
        fila_firmas = fila_resumen + 5
        ws[f'A{fila_firmas}'] = 'OBSERVACIONES:'
        ws[f'A{fila_firmas}'].font = header_font
        ws[f'A{fila_firmas}'].border = border_thin
        ws[f'A{fila_firmas}'].fill = light_gray
        
        ws.merge_cells(f'B{fila_firmas}:P{fila_firmas}')
        aplicar_borde_completo(ws, f'B{fila_firmas}:P{fila_firmas}', border_thin)
        
        # Firmas
        fila_firma = fila_firmas + 2
        ws[f'A{fila_firma}'] = 'Elaboró: Sergio Buitrago'
        ws[f'A{fila_firma}'].font = small_font
        ws[f'A{fila_firma}'].alignment = Alignment(horizontal='center', vertical='center')
        ws[f'A{fila_firmas}'].border = border_thin
        
        ws[f'M{fila_firma}'] = 'Administrador Unidades Agropecuarias'
        ws[f'M{fila_firma}'].font = small_font
        ws[f'M{fila_firma}'].alignment = Alignment(horizontal='right', vertical='center')
        ws[f'M{fila_firma}'].border = border_thin
        
        # Ajustar anchos de columna (actualizar para nueva estructura)
        column_widths = {
            'A': 21,   # Día
            'B': 19,  # 1a
            'C': 19,  # 2a
            'D': 19,  # 3a
            'E': 25,  # Rotos
            'F': 12,  # Total
            'G': 12,  # Promedio (nueva columna)
            'H': 12,  # Kg (movido)
            'I': 12,  # Acumulado (movido)
            'J': 10,  # Descartes (movido)
            'K': 10,  # Eliminación (movido)
            'L': 10,  # Total bajas (movido)
            'M': 12,  # Existencia (movido)
            'N': 15,  # Observaciones (movido)
            'O': 15,  # Observaciones
            'P': 15,  # Observaciones
        }
        
        for col, ancho in column_widths.items():
            ws.column_dimensions[col].width = ancho
        
        # Ajustar alturas de fila
        ws.row_dimensions[1].height = 25
        ws.row_dimensions[2].height = 25
        ws.row_dimensions[3].height = 25
        ws.row_dimensions[4].height = 20
        ws.row_dimensions[6].height = 20
        ws.row_dimensions[11].height = 20
        ws.row_dimensions[12].height = 20
        
        # Configurar respuesta HTTP
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
        lote_codigo = lote.codigo if lote else 'General'
        nombre_archivo = f'Registro_Postura_SENA_{lote_codigo}_{mes:02d}_{año}.xlsx'
        response['Content-Disposition'] = f'attachment; filename="{nombre_archivo}"'
        
        wb.save(response)
        return response
        
    except Exception as e:
        error_detail = traceback.format_exc()
        return HttpResponseServerError(f"Error al generar archivo Excel: {str(e)}\n{error_detail}")