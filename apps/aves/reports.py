"""
Sistema de reportes específicos para el módulo avícola
"""

import os
import io
import csv
from datetime import datetime, timedelta, date
from django.conf import settings
from django.template.loader import render_to_string
from django.http import HttpResponse, JsonResponse
from django.db.models import Sum, Avg, Count, Q, F, Max, Min
from django.utils import timezone
from django.contrib.auth.decorators import login_required
from django.shortcuts import get_object_or_404

# Importaciones para PDF y Excel
try:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import letter, A4
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch
    from reportlab.graphics.shapes import Drawing
    from reportlab.graphics.charts.linecharts import HorizontalLineChart
    from reportlab.graphics.charts.barcharts import VerticalBarChart
    REPORTLAB_AVAILABLE = True
except ImportError:
    REPORTLAB_AVAILABLE = False

try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.chart import BarChart, LineChart, Reference, PieChart
    from openpyxl.utils.dataframe import dataframe_to_rows
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

from .models import (
    LoteAves, BitacoraDiaria, MovimientoHuevos, ControlConcentrado,
    PlanVacunacion, AlertaSistema, TipoVacuna, TipoConcentrado
)

class ReporteAvicola:
    """
    Clase principal para generar reportes avícolas
    """
    
    def __init__(self, parametros=None):
        self.parametros = parametros or {}
        self.fecha_inicio = self.parametros.get('fecha_inicio')
        self.fecha_fin = self.parametros.get('fecha_fin')
        self.lote_id = self.parametros.get('lote_id')
        
    def obtener_datos_produccion_diaria(self):
        """
        Obtiene datos de producción diaria con filtros aplicados
        """
        queryset = BitacoraDiaria.objects.all()
        
        if self.lote_id:
            queryset = queryset.filter(lote_id=self.lote_id)
        if self.fecha_inicio:
            queryset = queryset.filter(fecha__gte=self.fecha_inicio)
        if self.fecha_fin:
            queryset = queryset.filter(fecha__lte=self.fecha_fin)
            
        return queryset.select_related('lote').order_by('-fecha')
    
    def obtener_resumen_produccion(self):
        """
        Obtiene resumen estadístico de producción
        """
        datos = self.obtener_datos_produccion_diaria()
        
        if not datos.exists():
            return {
                'total_huevos': 0,
                'produccion_aaa': 0,
                'produccion_aa': 0,
                'produccion_a': 0,
                'produccion_b': 0,
                'produccion_c': 0,
                'total_mortalidad': 0,
                'promedio_diario': 0,
                'mejor_dia': 0,
                'total_consumo': 0,
                'porcentaje_postura': 0,
                'dias_registrados': 0
            }
            
        resumen = datos.aggregate(
            total_huevos=Sum(F('produccion_aaa') + F('produccion_aa') + F('produccion_a') + F('produccion_b') + F('produccion_c')),
            produccion_aaa=Sum('produccion_aaa'),
            produccion_aa=Sum('produccion_aa'),
            produccion_a=Sum('produccion_a'),
            produccion_b=Sum('produccion_b'),
            produccion_c=Sum('produccion_c'),
            total_mortalidad=Sum('mortalidad'),
            promedio_diario=Avg(F('produccion_aaa') + F('produccion_aa') + F('produccion_a') + F('produccion_b') + F('produccion_c')),
            mejor_dia=Max(F('produccion_aaa') + F('produccion_aa') + F('produccion_a') + F('produccion_b') + F('produccion_c')),
            total_consumo=Sum('consumo_concentrado'),
            dias_registrados=Count('id')
        )
        
        # Calcular porcentaje de postura promedio
        total_aves = 0
        total_produccion = 0
        
        for registro in datos:
            aves_dia = registro.lote.numero_aves_actual
            produccion_dia = (registro.produccion_aaa + registro.produccion_aa + 
                            registro.produccion_a + registro.produccion_b + registro.produccion_c)
            total_aves += aves_dia
            total_produccion += produccion_dia
            
        porcentaje_postura = (total_produccion / total_aves * 100) if total_aves > 0 else 0
        
        # Calcular porcentajes por categoría
        total_huevos = resumen['total_huevos'] or 0
        porcentaje_aaa = (resumen['produccion_aaa'] / total_huevos * 100) if total_huevos > 0 else 0
        porcentaje_aa = (resumen['produccion_aa'] / total_huevos * 100) if total_huevos > 0 else 0
        porcentaje_a = (resumen['produccion_a'] / total_huevos * 100) if total_huevos > 0 else 0
        porcentaje_b = (resumen['produccion_b'] / total_huevos * 100) if total_huevos > 0 else 0
        porcentaje_c = (resumen['produccion_c'] / total_huevos * 100) if total_huevos > 0 else 0
        
        return {
            'total_huevos': total_huevos,
            'produccion_aaa': resumen['produccion_aaa'] or 0,
            'produccion_aa': resumen['produccion_aa'] or 0,
            'produccion_a': resumen['produccion_a'] or 0,
            'produccion_b': resumen['produccion_b'] or 0,
            'produccion_c': resumen['produccion_c'] or 0,
            'porcentaje_aaa': round(porcentaje_aaa, 2),
            'porcentaje_aa': round(porcentaje_aa, 2),
            'porcentaje_a': round(porcentaje_a, 2),
            'porcentaje_b': round(porcentaje_b, 2),
            'porcentaje_c': round(porcentaje_c, 2),
            'total_mortalidad': resumen['total_mortalidad'] or 0,
            'promedio_diario': round(resumen['promedio_diario'] or 0, 1),
            'mejor_dia': resumen['mejor_dia'] or 0,
            'total_consumo': resumen['total_consumo'] or 0,
            'porcentaje_postura': round(porcentaje_postura, 2),
            'dias_registrados': resumen['dias_registrados']
        }
    
    def obtener_datos_movimiento_huevos(self):
        """
        Obtiene datos de movimientos de huevos
        """
        queryset = MovimientoHuevos.objects.all()
        
        if self.fecha_inicio:
            queryset = queryset.filter(fecha__gte=self.fecha_inicio)
        if self.fecha_fin:
            queryset = queryset.filter(fecha__lte=self.fecha_fin)
            
        return queryset.prefetch_related('detalles').order_by('-fecha')
    
    def obtener_datos_consumo_concentrado(self):
        """
        Obtiene datos de consumo de concentrado
        """
        queryset = ControlConcentrado.objects.filter(tipo_movimiento='salida')
        
        if self.lote_id:
            queryset = queryset.filter(lote_id=self.lote_id)
        if self.fecha_inicio:
            queryset = queryset.filter(fecha__gte=self.fecha_inicio)
        if self.fecha_fin:
            queryset = queryset.filter(fecha__lte=self.fecha_fin)
            
        return queryset.select_related('tipo_concentrado', 'lote').order_by('-fecha')
    
    def obtener_datos_vacunacion(self):
        """
        Obtiene datos de vacunación
        """
        queryset = PlanVacunacion.objects.all()
        
        if self.lote_id:
            queryset = queryset.filter(lote_id=self.lote_id)
        if self.fecha_inicio:
            queryset = queryset.filter(fecha_programada__gte=self.fecha_inicio)
        if self.fecha_fin:
            queryset = queryset.filter(fecha_programada__lte=self.fecha_fin)
            
        return queryset.select_related('lote', 'tipo_vacuna', 'veterinario').order_by('-fecha_programada')
    
    def generar_pdf_produccion(self, nombre_archivo="reporte_produccion.pdf"):
        """
        Genera reporte PDF de producción
        """
        if not REPORTLAB_AVAILABLE:
            raise ImportError("ReportLab no está disponible")
            
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4)
        styles = getSampleStyleSheet()
        story = []
        
        # Título
        titulo_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=18,
            spaceAfter=30,
            alignment=1  # Centrado
        )
        story.append(Paragraph("Reporte de Producción Avícola", titulo_style))
        story.append(Spacer(1, 20))
        
        # Información del reporte
        info_data = [
            ['Fecha de generación:', datetime.now().strftime('%d/%m/%Y %H:%M')],
            ['Período:', f"{self.fecha_inicio or 'N/A'} - {self.fecha_fin or 'N/A'}"],
        ]
        info_table = Table(info_data, colWidths=[2*inch, 3*inch])
        info_table.setStyle(TableStyle([
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
        ]))
        story.append(info_table)
        story.append(Spacer(1, 20))
        
        # Resumen estadístico
        resumen = self.obtener_resumen_produccion()
        story.append(Paragraph("Resumen Estadístico", styles['Heading2']))
        
        resumen_data = [
            ['Concepto', 'Valor'],
            ['Total de huevos producidos', f"{resumen['total_huevos']:,}"],
            ['Producción AAA', f"{resumen['produccion_aaa']:,} ({resumen['porcentaje_aaa']}%)"],
            ['Producción AA', f"{resumen['produccion_aa']:,} ({resumen['porcentaje_aa']}%)"],
            ['Producción A', f"{resumen['produccion_a']:,} ({resumen['porcentaje_a']}%)"],
            ['Producción B', f"{resumen['produccion_b']:,} ({resumen['porcentaje_b']}%)"],
            ['Producción C', f"{resumen['produccion_c']:,} ({resumen['porcentaje_c']}%)"],
            ['Total mortalidad', f"{resumen['total_mortalidad']:,}"],
            ['Días registrados', f"{resumen['dias_registrados']}"],
            ['Promedio diario', f"{resumen['promedio_diario']:,}"],
            ['% Postura promedio', f"{resumen['porcentaje_postura']}%"],
        ]
        
        resumen_table = Table(resumen_data, colWidths=[3*inch, 2*inch])
        resumen_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 9),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        story.append(resumen_table)
        story.append(Spacer(1, 20))
        
        # Datos detallados de producción
        story.append(Paragraph("Detalle de Producción Diaria", styles['Heading2']))
        datos_produccion = self.obtener_datos_produccion_diaria()
        
        if datos_produccion:
            # Encabezados de la tabla
            headers = ['Fecha', 'Lote', 'Galpón', 'AAA', 'AA', 'A', 'B', 'C', 'Total', 'Mortalidad']
            data = [headers]
            
            for dato in datos_produccion[:50]:  # Limitar a 50 registros para el PDF
                total_huevos = dato.produccion_aaa + dato.produccion_aa + dato.produccion_a + dato.produccion_b + dato.produccion_c
                data.append([
                    dato.fecha.strftime('%d/%m/%Y'),
                    dato.lote.codigo,
                    dato.lote.galpon,
                    str(dato.produccion_aaa),
                    str(dato.produccion_aa),
                    str(dato.produccion_a),
                    str(dato.produccion_b),
                    str(dato.produccion_c),
                    str(total_huevos),
                    str(dato.mortalidad)
                ])
            
            produccion_table = Table(data, colWidths=[0.8*inch, 0.8*inch, 0.8*inch, 0.5*inch, 0.5*inch, 0.5*inch, 0.5*inch, 0.5*inch, 0.6*inch, 0.6*inch])
            produccion_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, -1), 8),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))
            story.append(produccion_table)
        
        doc.build(story)
        buffer.seek(0)
        
        response = HttpResponse(buffer.getvalue(), content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="{nombre_archivo}"'
        return response
    
    def generar_excel_produccion(self, nombre_archivo="reporte_produccion.xlsx"):
        """
        Genera reporte Excel de producción con múltiples hojas y gráficos
        """
        if not OPENPYXL_AVAILABLE:
            raise ImportError("openpyxl no está disponible")
            
        wb = openpyxl.Workbook()
        
        # Hoja 1: Resumen
        ws_resumen = wb.active
        ws_resumen.title = "Resumen"
        
        # Título
        ws_resumen['A1'] = "Reporte de Producción Avícola"
        ws_resumen['A1'].font = Font(size=16, bold=True)
        ws_resumen.merge_cells('A1:D1')
        
        # Información del reporte
        ws_resumen['A3'] = "Fecha de generación:"
        ws_resumen['B3'] = datetime.now().strftime('%d/%m/%Y %H:%M')
        ws_resumen['A4'] = "Período:"
        ws_resumen['B4'] = f"{self.fecha_inicio or 'N/A'} - {self.fecha_fin or 'N/A'}"
        
        # Resumen estadístico
        resumen = self.obtener_resumen_produccion()
        ws_resumen['A6'] = "Resumen Estadístico"
        ws_resumen['A6'].font = Font(size=14, bold=True)
        
        datos_resumen = [
            ['Concepto', 'Valor'],
            ['Total de huevos producidos', resumen['total_huevos']],
            ['Producción AAA', f"{resumen['produccion_aaa']} ({resumen['porcentaje_aaa']}%)"],
            ['Producción AA', f"{resumen['produccion_aa']} ({resumen['porcentaje_aa']}%)"],
            ['Producción A', f"{resumen['produccion_a']} ({resumen['porcentaje_a']}%)"],
            ['Producción B', f"{resumen['produccion_b']} ({resumen['porcentaje_b']}%)"],
            ['Producción C', f"{resumen['produccion_c']} ({resumen['porcentaje_c']}%)"],
            ['Total mortalidad', resumen['total_mortalidad']],
            ['Días registrados', resumen['dias_registrados']],
            ['Promedio diario', resumen['promedio_diario']],
            ['% Postura promedio', f"{resumen['porcentaje_postura']}%"],
        ]
        
        for i, fila in enumerate(datos_resumen, start=8):
            for j, valor in enumerate(fila, start=1):
                celda = ws_resumen.cell(row=i, column=j, value=valor)
                if i == 8:  # Encabezado
                    celda.font = Font(bold=True)
                    celda.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        
        # Hoja 2: Producción Diaria
        ws_produccion = wb.create_sheet("Producción Diaria")
        datos_produccion = self.obtener_datos_produccion_diaria()
        
        # Encabezados
        encabezados = ['Fecha', 'Lote', 'Galpón', 'Producción AAA', 'Producción AA', 'Producción A', 
                      'Producción B', 'Producción C', 'Total Huevos', 'Mortalidad', 'Consumo Concentrado', 'Observaciones']
        
        for i, encabezado in enumerate(encabezados, start=1):
            celda = ws_produccion.cell(row=1, column=i, value=encabezado)
            celda.font = Font(bold=True)
            celda.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        
        # Datos
        for i, dato in enumerate(datos_produccion, start=2):
            total_huevos = dato.produccion_aaa + dato.produccion_aa + dato.produccion_a + dato.produccion_b + dato.produccion_c
            ws_produccion.cell(row=i, column=1, value=dato.fecha)
            ws_produccion.cell(row=i, column=2, value=dato.lote.codigo)
            ws_produccion.cell(row=i, column=3, value=dato.lote.galpon)
            ws_produccion.cell(row=i, column=4, value=dato.produccion_aaa)
            ws_produccion.cell(row=i, column=5, value=dato.produccion_aa)
            ws_produccion.cell(row=i, column=6, value=dato.produccion_a)
            ws_produccion.cell(row=i, column=7, value=dato.produccion_b)
            ws_produccion.cell(row=i, column=8, value=dato.produccion_c)
            ws_produccion.cell(row=i, column=9, value=total_huevos)
            ws_produccion.cell(row=i, column=10, value=dato.mortalidad)
            ws_produccion.cell(row=i, column=11, value=float(dato.consumo_concentrado))
            ws_produccion.cell(row=i, column=12, value=dato.observaciones)
        
        # Ajustar ancho de columnas
        for column in ws_produccion.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws_produccion.column_dimensions[column_letter].width = adjusted_width
        
        # Crear gráfico de producción
        if datos_produccion:
            chart = LineChart()
            chart.title = "Evolución de la Producción de Huevos"
            chart.style = 13
            chart.x_axis.title = 'Fecha'
            chart.y_axis.title = 'Cantidad de Huevos'
            
            # Datos para el gráfico (últimos 30 días)
            datos_grafico = list(datos_produccion)[-30:] if len(list(datos_produccion)) > 30 else list(datos_produccion)
            
            # Agregar datos del gráfico en una nueva hoja
            ws_grafico = wb.create_sheet("Datos Gráfico")
            ws_grafico['A1'] = "Fecha"
            ws_grafico['B1'] = "Total Huevos"
            ws_grafico['C1'] = "Mortalidad"
            
            for i, dato in enumerate(datos_grafico, start=2):
                total_huevos = dato.produccion_aaa + dato.produccion_aa + dato.produccion_a + dato.produccion_b + dato.produccion_c
                ws_grafico.cell(row=i, column=1, value=dato.fecha.strftime('%d/%m'))
                ws_grafico.cell(row=i, column=2, value=total_huevos)
                ws_grafico.cell(row=i, column=3, value=dato.mortalidad)
            
            # Configurar referencias del gráfico
            data = Reference(ws_grafico, min_col=2, min_row=1, max_col=3, max_row=len(datos_grafico)+1)
            cats = Reference(ws_grafico, min_col=1, min_row=2, max_row=len(datos_grafico)+1)
            
            chart.add_data(data, titles_from_data=True)
            chart.set_categories(cats)
            
            # Agregar gráfico a la hoja de resumen
            ws_resumen.add_chart(chart, "F6")
        
        # Guardar en buffer
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        response = HttpResponse(
            buffer.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{nombre_archivo}"'
        return response
    
    def generar_csv_produccion(self, nombre_archivo="reporte_produccion.csv"):
        """
        Genera reporte CSV de producción
        """
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = f'attachment; filename="{nombre_archivo}"'
        
        writer = csv.writer(response)
        
        # Encabezados
        writer.writerow([
            'Fecha', 'Lote', 'Galpón', 'Producción AAA', 'Producción AA', 'Producción A',
            'Producción B', 'Producción C', 'Total Huevos', 'Mortalidad', 'Consumo Concentrado', 'Observaciones'
        ])
        
        # Datos
        datos_produccion = self.obtener_datos_produccion_diaria()
        for dato in datos_produccion:
            total_huevos = dato.produccion_aaa + dato.produccion_aa + dato.produccion_a + dato.produccion_b + dato.produccion_c
            writer.writerow([
                dato.fecha.strftime('%d/%m/%Y'),
                dato.lote.codigo,
                dato.lote.galpon,
                dato.produccion_aaa,
                dato.produccion_aa,
                dato.produccion_a,
                dato.produccion_b,
                dato.produccion_c,
                total_huevos,
                dato.mortalidad,
                float(dato.consumo_concentrado),
                dato.observaciones
            ])
        
        return response

class ReporteComparativo:
    """
    Clase para generar reportes comparativos entre lotes, períodos, etc.
    """
    
    def __init__(self, parametros=None):
        self.parametros = parametros or {}
        
    def comparar_lotes(self, lotes_ids, fecha_inicio, fecha_fin):
        """
        Compara producción entre diferentes lotes
        """
        datos_comparacion = []
        
        for lote_id in lotes_ids:
            lote = get_object_or_404(LoteAves, id=lote_id)
            
            # Obtener datos del lote
            bitacoras = BitacoraDiaria.objects.filter(
                lote_id=lote_id,
                fecha__gte=fecha_inicio,
                fecha__lte=fecha_fin
            )
            
            resumen = bitacoras.aggregate(
                total_produccion_aaa=Sum('produccion_aaa'),
                total_produccion_aa=Sum('produccion_aa'),
                total_produccion_a=Sum('produccion_a'),
                total_produccion_b=Sum('produccion_b'),
                total_produccion_c=Sum('produccion_c'),
                total_mortalidad=Sum('mortalidad'),
                promedio_postura=Avg(F('produccion_aaa') + F('produccion_aa') + F('produccion_a') + F('produccion_b') + F('produccion_c')),
                dias_registrados=Count('id')
            )
            
            total_huevos = (
                (resumen['total_produccion_aaa'] or 0) + 
                (resumen['total_produccion_aa'] or 0) + 
                (resumen['total_produccion_a'] or 0) + 
                (resumen['total_produccion_b'] or 0) + 
                (resumen['total_produccion_c'] or 0)
            )
            
            datos_comparacion.append({
                'lote_codigo': lote.codigo,
                'galpon': lote.galpon,
                'linea_genetica': lote.linea_genetica,
                'numero_aves_actual': lote.numero_aves_actual,
                'total_huevos': total_huevos,
                'produccion_aaa': resumen['total_produccion_aaa'] or 0,
                'produccion_aa': resumen['total_produccion_aa'] or 0,
                'produccion_a': resumen['total_produccion_a'] or 0,
                'produccion_b': resumen['total_produccion_b'] or 0,
                'produccion_c': resumen['total_produccion_c'] or 0,
                'total_mortalidad': resumen['total_mortalidad'] or 0,
                'promedio_postura': round(resumen['promedio_postura'] or 0, 2),
                'dias_registrados': resumen['dias_registrados'],
                'huevos_por_ave_dia': round(total_huevos / (lote.numero_aves_actual * resumen['dias_registrados']), 3) if lote.numero_aves_actual > 0 and resumen['dias_registrados'] > 0 else 0
            })
        
        return datos_comparacion
    
    def comparar_periodos(self, lote_id, periodos):
        """
        Compara producción del mismo lote en diferentes períodos
        """
        datos_comparacion = []
        
        for periodo in periodos:
            fecha_inicio = periodo['fecha_inicio']
            fecha_fin = periodo['fecha_fin']
            nombre_periodo = periodo['nombre']
            
            bitacoras = BitacoraDiaria.objects.filter(
                lote_id=lote_id,
                fecha__gte=fecha_inicio,
                fecha__lte=fecha_fin
            )
            
            resumen = bitacoras.aggregate(
                total_produccion_aaa=Sum('produccion_aaa'),
                total_produccion_aa=Sum('produccion_aa'),
                total_produccion_a=Sum('produccion_a'),
                total_produccion_b=Sum('produccion_b'),
                total_produccion_c=Sum('produccion_c'),
                total_mortalidad=Sum('mortalidad'),
                dias_registrados=Count('id')
            )
            
            total_huevos = (
                (resumen['total_produccion_aaa'] or 0) + 
                (resumen['total_produccion_aa'] or 0) + 
                (resumen['total_produccion_a'] or 0) + 
                (resumen['total_produccion_b'] or 0) + 
                (resumen['total_produccion_c'] or 0)
            )
            
            datos_comparacion.append({
                'periodo': nombre_periodo,
                'fecha_inicio': fecha_inicio,
                'fecha_fin': fecha_fin,
                'total_huevos': total_huevos,
                'produccion_aaa': resumen['total_produccion_aaa'] or 0,
                'produccion_aa': resumen['total_produccion_aa'] or 0,
                'produccion_a': resumen['total_produccion_a'] or 0,
                'produccion_b': resumen['total_produccion_b'] or 0,
                'produccion_c': resumen['total_produccion_c'] or 0,
                'total_mortalidad': resumen['total_mortalidad'] or 0,
                'dias_registrados': resumen['dias_registrados'],
                'promedio_huevos_dia': round(total_huevos / resumen['dias_registrados'], 1) if resumen['dias_registrados'] > 0 else 0
            })
        
        return datos_comparacion

def obtener_datos_dashboard():
    """
    Obtiene datos para el dashboard principal del módulo avícola
    """
    hoy = timezone.now().date()
    hace_30_dias = hoy - timedelta(days=30)
    
    # Estadísticas generales
    total_lotes_activos = LoteAves.objects.filter(estado__in=['levante', 'postura']).count()
    total_aves = LoteAves.objects.filter(estado__in=['levante', 'postura']).aggregate(total=Sum('numero_aves_actual'))['total'] or 0
    
    # Producción del día
    produccion_hoy = BitacoraDiaria.objects.filter(fecha=hoy).aggregate(
        huevos_hoy=Sum(F('produccion_aaa') + F('produccion_aa') + F('produccion_a') + F('produccion_b') + F('produccion_c')),
        mortalidad_hoy=Sum('mortalidad')
    )
    
    # Producción últimos 30 días
    produccion_30_dias = BitacoraDiaria.objects.filter(
        fecha__gte=hace_30_dias,
        fecha__lte=hoy
    ).aggregate(
        total_huevos=Sum(F('produccion_aaa') + F('produccion_aa') + F('produccion_a') + F('produccion_b') + F('produccion_c')),
        promedio_diario=Avg(F('produccion_aaa') + F('produccion_aa') + F('produccion_a') + F('produccion_b') + F('produccion_c'))
    )
    
    # Alertas activas
    alertas_activas = AlertaSistema.objects.filter(
        leida=False
    ).count()
    
    # Evolución de producción (últimos 7 días)
    evolucion_produccion = []
    for i in range(7):
        fecha = hoy - timedelta(days=i)
        produccion_dia = BitacoraDiaria.objects.filter(fecha=fecha).aggregate(
            total=Sum(F('produccion_aaa') + F('produccion_aa') + F('produccion_a') + F('produccion_b') + F('produccion_c'))
        )['total'] or 0
        evolucion_produccion.append({
            'fecha': fecha.strftime('%d/%m'),
            'produccion': produccion_dia
        })
    
    evolucion_produccion.reverse()
    
    # Top 5 lotes por producción
    top_lotes = BitacoraDiaria.objects.filter(
        fecha__gte=hace_30_dias
    ).values(
        'lote__codigo', 'lote__galpon'
    ).annotate(
        total_produccion=Sum(F('produccion_aaa') + F('produccion_aa') + F('produccion_a') + F('produccion_b') + F('produccion_c'))
    ).order_by('-total_produccion')[:5]
    
    return {
        'total_lotes_activos': total_lotes_activos,
        'total_aves': total_aves,
        'huevos_hoy': produccion_hoy['huevos_hoy'] or 0,
        'mortalidad_hoy': produccion_hoy['mortalidad_hoy'] or 0,
        'total_huevos_30_dias': produccion_30_dias['total_huevos'] or 0,
        'promedio_diario_30_dias': round(produccion_30_dias['promedio_diario'] or 0, 1),
        'alertas_activas': alertas_activas,
        'evolucion_produccion': evolucion_produccion,
        'top_lotes': list(top_lotes)
    }


def generar_reporte_sena_excel(lote_id, mes, año, nombre_granja="Granja Avícola La Salada", registro_ica="051290274"):
    """
    Genera reporte mensual en formato SENA para registro de postura
    """
    if not OPENPYXL_AVAILABLE:
        raise ImportError("openpyxl no está disponible para generar reportes Excel")
    
    # Obtener el lote
    try:
        lote = LoteAves.objects.get(id=lote_id)
    except LoteAves.DoesNotExist:
        raise ValueError(f"No se encontró el lote con ID {lote_id}")
    
    # Crear workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Registro Postura {mes}-{año}"
    
    # Configurar página
    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    
    # Estilos
    header_font = Font(name='Arial', size=10, bold=True)
    normal_font = Font(name='Arial', size=9)
    title_font = Font(name='Arial', size=12, bold=True)
    border_thin = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Colores
    green_fill = PatternFill(start_color='92D050', end_color='92D050', fill_type='solid')
    light_green_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
    
    # ENCABEZADO SENA
    # Fila 1: Logo SENA, Nombre de la granja, Centro
    ws['A1'] = 'SENA'
    ws['A1'].font = Font(name='Arial', size=14, bold=True)
    ws['A1'].fill = green_fill
    
    ws.merge_cells('B1:M1')
    ws['B1'] = nombre_granja.upper()
    ws['B1'].font = title_font
    ws['B1'].alignment = Alignment(horizontal='center')
    ws['B1'].fill = green_fill
    
    ws.merge_cells('N1:P1')
    ws['N1'] = 'CENTRO DE LOS RECURSOS NATURALES RENOVABLES'
    ws['N1'].font = Font(name='Arial', size=8)
    ws['N1'].alignment = Alignment(horizontal='center', wrap_text=True)
    ws['N1'].fill = green_fill
    
    # Fila 2: Registro ICA
    ws.merge_cells('A2:P2')
    ws['A2'] = f'Registro ICA {registro_ica}'
    ws['A2'].font = Font(name='Arial', size=11, bold=True)
    ws['A2'].alignment = Alignment(horizontal='center')
    ws['A2'].fill = light_green_fill
    
    # Fila 3: Vacía
    
    # Fila 4: Título del reporte
    ws.merge_cells('A4:P4')
    ws['A4'] = 'REGISTRO MENSUAL DE POSTURA'
    ws['A4'].font = Font(name='Arial', size=14, bold=True)
    ws['A4'].alignment = Alignment(horizontal='center')
    ws['A4'].fill = green_fill
    
    # Fila 5: Versión
    ws['P5'] = 'Versión: 2020-01'
    ws['P5'].font = Font(name='Arial', size=8)
    ws['P5'].alignment = Alignment(horizontal='right')
    
    # Fila 6: Información del lote
    ws['A6'] = 'Mes:'
    ws['A6'].font = header_font
    ws['B6'] = f'{mes:02d}'
    ws['B6'].font = normal_font
    
    ws['C6'] = 'Año:'
    ws['C6'].font = header_font
    ws['D6'] = str(año)
    ws['D6'].font = normal_font
    
    ws['E6'] = 'Galpón:'
    ws['E6'].font = header_font
    ws['F6'] = str(lote.galpon)
    ws['F6'].font = normal_font
    
    ws['G6'] = 'Lote:'
    ws['G6'].font = header_font
    ws['H6'] = str(lote.codigo)
    ws['H6'].font = normal_font
    
    ws['I6'] = 'Sistema:'
    ws['I6'].font = header_font
    ws['J6'] = 'Jaula'  # Valor por defecto
    ws['J6'].font = normal_font
    
    ws['K6'] = 'Línea:'
    ws['K6'].font = header_font
    ws['L6'] = lote.get_linea_genetica_display_name() if hasattr(lote, 'get_linea_genetica_display_name') else str(lote.linea_genetica)
    ws['L6'].font = normal_font
    
    # Fila 7: Información de aves
    ws['A7'] = 'Edad en semanas:'
    ws['A7'].font = header_font
    edad_semanas = lote.edad_actual_semanas if hasattr(lote, 'edad_actual_semanas') else 0
    ws['B7'] = f'{edad_semanas:.1f}'
    ws['B7'].font = normal_font
    
    ws['C7'] = 'Aves alojadas:'
    ws['C7'].font = header_font
    ws['D7'] = lote.numero_aves_inicial
    ws['D7'].font = normal_font
    
    ws['E7'] = 'N° de aves al inicio del mes:'
    ws['E7'].font = header_font
    ws['F7'] = lote.numero_aves_actual
    ws['F7'].font = normal_font
    
    # TABLA DE DATOS DIARIOS
    # Fila 9: Encabezados principales
    fila_encabezado = 9
    
    # Aplicar color de fondo a encabezados
    for col in range(1, 17):  # A hasta P
        ws.cell(row=fila_encabezado, column=col).fill = green_fill
        ws.cell(row=fila_encabezado + 1, column=col).fill = light_green_fill
    
    # Encabezados principales
    ws['A9'] = 'Día'
    ws.merge_cells('B9:F9')
    ws['B9'] = 'PRODUCCIÓN'
    ws.merge_cells('G9:H9')
    ws['G9'] = 'ALIMENTO'
    ws.merge_cells('I9:K9')
    ws['I9'] = 'BAJAS'
    ws['L9'] = 'Existencia'
    ws['M9'] = 'OBSERVACIONES'
    
    # Aplicar formato a encabezados principales
    for col in ['A', 'B', 'G', 'I', 'L', 'M']:
        ws[f'{col}9'].font = header_font
        ws[f'{col}9'].alignment = Alignment(horizontal='center', vertical='center')
        ws[f'{col}9'].border = border_thin
    
    # Fila 10: Subencabezados
    subencabezados = [
        ('A', ''),
        ('B', '1a'),
        ('C', '2a'),
        ('D', '3a'),
        ('E', 'Rotos'),
        ('F', 'Total'),
        ('G', 'Kg'),
        ('H', 'Acumulado'),
        ('I', 'Descar'),
        ('J', 'Elimin'),
        ('K', 'Total'),
        ('L', ''),
        ('M', '')
    ]
    
    for col, texto in subencabezados:
        ws[f'{col}10'] = texto
        ws[f'{col}10'].font = header_font
        ws[f'{col}10'].alignment = Alignment(horizontal='center', vertical='center')
        ws[f'{col}10'].border = border_thin
    
    # Obtener datos del mes
    from calendar import monthrange
    dias_en_mes = monthrange(año, mes)[1]
    
    # Obtener bitácoras del mes
    bitacoras = BitacoraDiaria.objects.filter(
        lote=lote,
        fecha__year=año,
        fecha__month=mes
    ).order_by('fecha')
    
    # Crear diccionario de bitácoras por día
    bitacoras_por_dia = {bitacora.fecha.day: bitacora for bitacora in bitacoras}
    
    # LLENAR DATOS DIARIOS
    fila_inicio_datos = 11
    acumulado_alimento = 0
    
    # Variables para totales
    total_1a = 0
    total_2a = 0
    total_3a = 0
    total_rotos = 0
    total_alimento = 0
    total_mortalidad = 0
    
    for dia in range(1, dias_en_mes + 1):
        fila = fila_inicio_datos + dia - 1
        
        # Día
        ws[f'A{fila}'] = dia
        ws[f'A{fila}'].alignment = Alignment(horizontal='center')
        ws[f'A{fila}'].border = border_thin
        
        if dia in bitacoras_por_dia:
            bitacora = bitacoras_por_dia[dia]
            
            # Producción (mapear categorías del sistema a formato SENA)
            # 1a = AAA + AA
            # 2a = A
            # 3a = B + C
            produccion_1a = (bitacora.produccion_aaa or 0) + (bitacora.produccion_aa or 0)
            produccion_2a = bitacora.produccion_a or 0
            produccion_3a = (bitacora.produccion_b or 0) + (bitacora.produccion_c or 0)
            rotos = bitacora.huevos_rotos or 0
            total_produccion_dia = produccion_1a + produccion_2a + produccion_3a + rotos
            
            # Acumular totales
            total_1a += produccion_1a
            total_2a += produccion_2a
            total_3a += produccion_3a
            total_rotos += rotos
            
            ws[f'B{fila}'] = produccion_1a if produccion_1a > 0 else ''
            ws[f'C{fila}'] = produccion_2a if produccion_2a > 0 else ''
            ws[f'D{fila}'] = produccion_3a if produccion_3a > 0 else ''
            ws[f'E{fila}'] = rotos if rotos > 0 else ''
            ws[f'F{fila}'] = total_produccion_dia if total_produccion_dia > 0 else ''
            
            # Alimento
            consumo_kg = float(bitacora.consumo_concentrado) if bitacora.consumo_concentrado else 0
            acumulado_alimento += consumo_kg
            total_alimento += consumo_kg
            
            ws[f'G{fila}'] = consumo_kg if consumo_kg > 0 else ''
            ws[f'H{fila}'] = round(acumulado_alimento, 2) if acumulado_alimento > 0 else ''
            
            # Bajas (mortalidad)
            mortalidad = bitacora.mortalidad or 0
            total_mortalidad += mortalidad
            
            ws[f'I{fila}'] = mortalidad if mortalidad > 0 else ''
            ws[f'J{fila}'] = ''  # Eliminación (no tenemos este dato)
            ws[f'K{fila}'] = mortalidad if mortalidad > 0 else ''
            
            # Existencia (aves restantes)
            aves_restantes = lote.numero_aves_actual - total_mortalidad
            ws[f'L{fila}'] = aves_restantes if aves_restantes >= 0 else ''
            
            # Observaciones
            ws[f'M{fila}'] = bitacora.observaciones[:30] if bitacora.observaciones else ''
        
        # Aplicar bordes y formato a toda la fila
        for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M']:
            ws[f'{col}{fila}'].border = border_thin
            ws[f'{col}{fila}'].font = normal_font
            if col in ['B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L']:
                ws[f'{col}{fila}'].alignment = Alignment(horizontal='center')
    
    # FILA TOTAL
    fila_total = fila_inicio_datos + dias_en_mes
    ws[f'A{fila_total}'] = 'TOTAL'
    ws[f'A{fila_total}'].font = header_font
    ws[f'A{fila_total}'].fill = green_fill
    
    total_general = total_1a + total_2a + total_3a + total_rotos
    
    ws[f'B{fila_total}'] = total_1a if total_1a > 0 else ''
    ws[f'C{fila_total}'] = total_2a if total_2a > 0 else ''
    ws[f'D{fila_total}'] = total_3a if total_3a > 0 else ''
    ws[f'E{fila_total}'] = total_rotos if total_rotos > 0 else ''
    ws[f'F{fila_total}'] = total_general if total_general > 0 else ''
    
    # Total alimento
    ws[f'G{fila_total}'] = round(total_alimento, 2) if total_alimento > 0 else ''
    
    # Total bajas
    ws[f'I{fila_total}'] = total_mortalidad if total_mortalidad > 0 else ''
    ws[f'K{fila_total}'] = total_mortalidad if total_mortalidad > 0 else ''
    
    # Aplicar formato a fila total
    for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M']:
        ws[f'{col}{fila_total}'].border = border_thin
        ws[f'{col}{fila_total}'].font = header_font
        ws[f'{col}{fila_total}'].fill = green_fill
        if col in ['B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L']:
            ws[f'{col}{fila_total}'].alignment = Alignment(horizontal='center')
    
    # RESUMEN MENSUAL
    fila_resumen = fila_total + 3
    ws[f'A{fila_resumen}'] = 'RESUMEN MENSUAL'
    ws[f'A{fila_resumen}'].font = Font(name='Arial', size=12, bold=True)
    ws[f'A{fila_resumen}'].fill = green_fill
    
    fila_resumen += 1
    ws[f'A{fila_resumen}'] = 'PRODUCCIÓN TOTAL:'
    ws[f'A{fila_resumen}'].font = header_font
    ws[f'C{fila_resumen}'] = f'{total_general} huevos'
    ws[f'C{fila_resumen}'].font = normal_font
    
    ws[f'F{fila_resumen}'] = '% DE PRODUCCIÓN:'
    ws[f'F{fila_resumen}'].font = header_font
    
    # Calcular porcentaje de producción
    if lote.numero_aves_actual > 0:
        porcentaje_produccion = (total_general / (lote.numero_aves_actual * dias_en_mes)) * 100
        ws[f'H{fila_resumen}'] = f'{porcentaje_produccion:.1f}%'
        ws[f'H{fila_resumen}'].font = normal_font
    
    fila_resumen += 1
    ws[f'A{fila_resumen}'] = 'CONSUMO TOTAL:'
    ws[f'A{fila_resumen}'].font = header_font
    ws[f'C{fila_resumen}'] = f'{total_alimento:.2f} kg'
    ws[f'C{fila_resumen}'].font = normal_font
    
    ws[f'F{fila_resumen}'] = '% DE MORTALIDAD:'
    ws[f'F{fila_resumen}'].font = header_font
    
    # Calcular porcentaje de mortalidad
    if lote.numero_aves_inicial > 0:
        porcentaje_mortalidad = (total_mortalidad / lote.numero_aves_inicial) * 100
        ws[f'H{fila_resumen}'] = f'{porcentaje_mortalidad:.2f}%'
        ws[f'H{fila_resumen}'].font = normal_font
    
    fila_resumen += 1
    ws[f'A{fila_resumen}'] = 'MORTALIDAD TOTAL:'
    ws[f'A{fila_resumen}'].font = header_font
    ws[f'C{fila_resumen}'] = f'{total_mortalidad} aves'
    ws[f'C{fila_resumen}'].font = normal_font
    
    ws[f'F{fila_resumen}'] = 'CONSUMO PROMEDIO:'
    ws[f'F{fila_resumen}'].font = header_font
    if dias_en_mes > 0:
        consumo_promedio = total_alimento / dias_en_mes
        ws[f'H{fila_resumen}'] = f'{consumo_promedio:.2f} kg/día'
        ws[f'H{fila_resumen}'].font = normal_font
    
    # CONVERSIÓN Y FIRMAS
    fila_conversion = fila_resumen + 3
    ws[f'A{fila_conversion}'] = 'CONVERSIÓN:'
    ws[f'A{fila_conversion}'].font = header_font
    
    # Calcular conversión alimenticia
    if total_general > 0:
        # Conversión = kg alimento / docenas de huevos
        docenas = total_general / 12
        conversion = total_alimento / docenas if docenas > 0 else 0
        ws[f'C{fila_conversion}'] = f'{conversion:.2f} kg/docena'
        ws[f'C{fila_conversion}'].font = normal_font
    
    # Firmas
    fila_firmas = fila_conversion + 3
    ws[f'B{fila_firmas}'] = 'Elaboró: Sergio Buitrago'
    ws[f'B{fila_firmas}'].font = normal_font
    
    ws[f'H{fila_firmas}'] = 'Administrador Unidades Agropecuarias'
    ws[f'H{fila_firmas}'].font = normal_font
    
    # Ajustar ancho de columnas
    anchos_columnas = {
        'A': 6, 'B': 8, 'C': 8, 'D': 8, 'E': 8, 'F': 8,
        'G': 10, 'H': 12, 'I': 8, 'J': 8, 'K': 8, 'L': 10, 'M': 25
    }
    
    for col, ancho in anchos_columnas.items():
        ws.column_dimensions[col].width = ancho
    
    # Ajustar altura de filas
    ws.row_dimensions[1].height = 25
    ws.row_dimensions[2].height = 20
    ws.row_dimensions[4].height = 25
    ws.row_dimensions[9].height = 20
    ws.row_dimensions[10].height = 20
    
    # Crear respuesta HTTP
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    
    nombre_archivo = f'Registro_Postura_SENA_{lote.codigo}_{mes:02d}_{año}.xlsx'
    response['Content-Disposition'] = f'attachment; filename="{nombre_archivo}"'
    
    wb.save(response)
    return response